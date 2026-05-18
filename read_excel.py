import openpyxl
import os
import json

excel_path = "QuestionBank.xlsm"

if not os.path.exists(excel_path):
    print(f"File not found: {excel_path}")
    exit(1)

try:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    result = {}
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        
        # Parse rows
        sheet_data = []
        if len(rows) > 0:
            headers = [str(h).strip() if h is not None else f"Column_{i}" for i, h in enumerate(rows[0])]
            for r_idx, row in enumerate(rows[1:]):
                # Filter out entirely empty rows
                if not any(cell is not None for cell in row):
                    continue
                row_dict = {}
                for c_idx, cell in enumerate(row):
                    if c_idx < len(headers):
                        row_dict[headers[c_idx]] = cell
                sheet_data.append(row_dict)
        
        result[sheet_name] = sheet_data
        
    print(json.dumps({
        "sheets": wb.sheetnames,
        "first_sheet_preview": list(result.values())[0][:2] if result else []
    }, indent=2))
    
    # Save the parsed data to a JSON file in the React project
    os.makedirs("src/data", exist_ok=True)
    with open("src/data/questions.json", "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2, default=str)
    print("Successfully converted and saved to src/data/questions.json")

except Exception as e:
    print("Error reading workbook:", e)
